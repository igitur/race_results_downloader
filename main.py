import argparse

import pandas as pd
from loguru import logger
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from scrapers.scraper_factory import get_scraper

logger.add("log.txt", rotation="500 MB", level="DEBUG")


def main():
    parser = argparse.ArgumentParser(
        description="A utility to extract results from finishtime.co.za and dump it to a file."
    )

    parser.add_argument("-u", "--url", help="Exact full url on finishtime")
    parser.add_argument(
        "-o",
        "--output",
        type=str,
        default="results.csv",
        help="Output file (default: results.csv)",
    )

    args = parser.parse_args()

    if args.url is not None:
        scraper = get_scraper(args.url)
        results = scraper.get_results()
        _export_results(results, args.output)


def _export_results(results: list, output_filename: str):
    if len(results) == 0:
        logger.error("No results to export")
        return

    df = pd.DataFrame(results)

    df.drop(columns=["Fav", "Share", "Behind", ""], inplace=True, errors="ignore")

    # Ensure RaceName and EventName are the first 2 columns
    if all(col in df.columns for col in ["RaceName", "EventName"]):
        df = df[
            ["RaceName", "EventName"]
            + [col for col in df.columns if col not in ["RaceName", "EventName"]]
        ]

    file_extension = output_filename.split(".")[-1].lower()
    if file_extension == "csv":
        df.to_csv(output_filename, index=False)
        return

    numeric_columns = ["Pos", "CatPos", "GenPos", "GenPts", "Rank"]
    datetime_columns = ["StartTime"]
    duration_columns = ["Time", "Finish", "ResultTime", "Pace"]

    for nc in numeric_columns:
        if nc in df.columns:
            df[nc] = df[nc].apply(lambda x: int(x) if x.isnumeric() else x)

    for dtc in datetime_columns:
        if dtc in df.columns:
            df[dtc] = pd.to_datetime(df[dtc], errors="coerce").fillna(df[dtc])

    for dc in duration_columns:
        if dc in df.columns:
            df[dc] = pd.to_timedelta(df[dc], errors="coerce").fillna(df[dc])

    if file_extension == "xlsx":
        with pd.ExcelWriter(output_filename, engine="openpyxl") as writer:  # pylint: disable=abstract-class-instantiated
            df.to_excel(writer, sheet_name="Results", index=False)

            # Access the underlying openpyxl workbook and worksheet objects
            wb: Workbook = writer.book
            ws: Worksheet = wb.active

            header_row = next(ws.rows)
            for index, header_cell in enumerate(header_row):
                if header_cell.value in duration_columns:
                    col = ws.column_dimensions[get_column_letter(index + 1)]
                    col.number_format = "hh:mm:ss.000"
                    for value_cell in ws[get_column_letter(index + 1)]:
                        value_cell.number_format = "hh:mm:ss.000"

            _auto_size_columns(ws)

    elif file_extension == "json":
        df.to_json(output_filename, orient="records")
    else:
        print(f"Unsupported file format: {file_extension}")


def _auto_size_columns(ws: Worksheet):
    # Iterate over all columns and adjust their widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception as e:
                logger.warning(f"Could not get length of cell: {cell.coordinate} - {e}")

        adjusted_width = (max_length + 2) * 1.1
        ws.column_dimensions[column_letter].width = adjusted_width


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        logger.error(f"{e}")
        raise e
